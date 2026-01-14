#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
고급 트래픽 데이터 분석 대시보드 (개선 버전 v2)
- B1 셀에서 주차 자동 인식
- 헤더 행 자동 감지 (1행 또는 2행)
- 중복 인덱스 자동 제거로 Reindexing 오류 완전 해결
- 주차별 필터링 및 기간 비교
- 동적 컬럼 선택 및 자동 차트 생성
- '의견' 컬럼 상단 알림 박스 표시
"""

import os
import io
import base64
import pandas as pd
import openpyxl
from dash import Dash, dcc, html, dash_table, Input, Output, State
import dash_bootstrap_components as dbc
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ============================================================================
# 헬퍼 함수: Excel 파일 파싱
# ============================================================================

def read_excel_with_week_improved(file_content):
    """
    개선된 Excel 파싱 함수
    - B1 셀에서 주차 정보 안전하게 읽기
    - 헤더 자동 감지 (1행 또는 2행)
    - 중복 인덱스 자동 제거
    - 빈 행 필터링
    - 에러 핸들링 강화
    """
    all_data = []
    errors = []

    try:
        # BytesIO 객체 생성
        if isinstance(file_content, str):
            # base64 디코딩
            file_bytes = base64.b64decode(file_content.split(',')[1])
            excel_file = io.BytesIO(file_bytes)
        else:
            excel_file = io.BytesIO(file_content)

        # openpyxl로 워크북 로드
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            try:
                ws = wb[sheet_name]

                # B1 셀에서 주차 정보 읽기
                b1_value = ws['B1'].value
                week_info = str(b1_value) if b1_value else sheet_name

                # 헤더 행 자동 감지
                row1_values = [cell.value for cell in ws[1]]
                row2_values = [cell.value for cell in ws[2]] if ws.max_row >= 2 else []

                if '상품명' in row1_values or 'Product' in row1_values:
                    header_row = 0
                elif '상품명' in row2_values or 'Product' in row2_values:
                    header_row = 1
                else:
                    header_row = 1  # 기본값

                # pandas로 데이터 읽기
                excel_file.seek(0)
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)

                # 빈 행 제거
                df = df.dropna(how='all')

                # 컬럼명 표준화
                column_aliases = {
                    '상품명': ['상품명', 'Product', '제품명'],
                    '주차': ['주차', 'Week'],
                    '쇼핑몰': ['쇼핑몰', 'Mall', 'Shop'],
                    '매출': ['매출', 'Sales', '판매액'],
                    '이익액': ['이익액', 'Profit', '이익'],
                    '트래픽 비용': ['트래픽 비용', 'Traffic Cost', '비용'],
                    '이익액-비용': ['이익액-비용', 'Net Profit'],
                    '이익률(ROI)': ['이익률(ROI)', 'ROI', '이익률', 'Margin'],
                    '이익률 변동': ['이익률 변동', 'ROI Change'],
                    '슬롯수': ['슬롯수', 'Slots'],
                    '의견': ['의견', 'Opinion', 'Comment', 'Note']
                }

                col_mapping = {}
                for standard_name, aliases in column_aliases.items():
                    for col in df.columns:
                        if any(alias in str(col) for alias in aliases):
                            col_mapping[col] = standard_name
                            break

                if col_mapping:
                    df = df.rename(columns=col_mapping)

                # 주차 정보 추가
                df['시트명'] = sheet_name
                df['주차(B1)'] = week_info

                # 인덱스 리셋 (중요!)
                df = df.reset_index(drop=True)

                # 중복 행 제거
                if '상품명' in df.columns:
                    df = df.drop_duplicates(subset=['상품명', '주차(B1)'], keep='last')

                all_data.append(df)

            except Exception as e:
                errors.append(f"시트 '{sheet_name}' 처리 실패: {str(e)}")
                continue

        wb.close()

        # 모든 시트 데이터 병합
        if all_data:
            combined_df = pd.concat(all_data, ignore_index=True)
            combined_df = combined_df.reset_index(drop=True)  # 최종 인덱스 리셋
            return combined_df, errors
        else:
            return None, errors + ["모든 시트에서 데이터 로드 실패"]

    except Exception as e:
        return None, [f"파일 읽기 실패: {str(e)}"]

# ============================================================================
# Dash 앱 초기화
# ============================================================================

app = Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
server = app.server

# ============================================================================
# 레이아웃 정의
# ============================================================================

app.layout = dbc.Container([
    # 타이틀
    dbc.Row([
        dbc.Col([
            html.H1("📊 LABONLAB 트래픽 데이터 분석 대시보드", className="text-center text-primary mb-4")
        ])
    ]),

    # 의견 알림 박스 (M1/K 컬럼)
    dbc.Row([
        dbc.Col([
            dbc.Alert(
                id='opinion-alert',
                children="📢 의견 알림: 파일을 업로드하면 의견 내용이 여기에 표시됩니다.",
                color="info",
                dismissable=False,
                className="mb-4"
            )
        ])
    ]),

    # 파일 업로드
    dbc.Row([
        dbc.Col([
            dcc.Upload(
                id='upload-data',
                children=html.Div([
                    '📤 Excel 파일을 드래그 앤 드롭하거나 ',
                    html.A('클릭하여 업로드', style={'color': '#007bff', 'fontWeight': 'bold'})
                ]),
                style={
                    'width': '100%',
                    'height': '80px',
                    'lineHeight': '80px',
                    'borderWidth': '2px',
                    'borderStyle': 'dashed',
                    'borderRadius': '10px',
                    'borderColor': '#007bff',
                    'textAlign': 'center',
                    'backgroundColor': '#f8f9fa',
                    'cursor': 'pointer'
                },
                multiple=False
            ),
            html.Div(id='upload-status', className='mt-2')
        ])
    ], className='mb-4'),

    # 필터 영역
    dbc.Row([
        dbc.Col([
            html.Label("🗓️ 주차 선택:", className="fw-bold"),
            dcc.Dropdown(
                id='week-dropdown',
                options=[],
                value=None,
                placeholder="주차를 선택하세요",
                clearable=False
            )
        ], width=3),
        dbc.Col([
            html.Label("📊 비교 주차 선택:", className="fw-bold"),
            dcc.Dropdown(
                id='compare-week-dropdown',
                options=[],
                value=None,
                placeholder="비교할 주차 선택 (선택 사항)",
                clearable=True
            )
        ], width=3),
        dbc.Col([
            html.Label("📋 표시할 컬럼 선택:", className="fw-bold"),
            dcc.Checklist(
                id='column-checklist',
                options=[],
                value=[],
                inline=True,
                className="mt-2",
                style={'fontSize': '14px'}
            )
        ], width=6)
    ], className='mb-4'),

    # KPI 카드
    dbc.Row(id='kpi-cards', className='mb-4'),

    # 데이터 테이블
    dbc.Row([
        dbc.Col([
            html.H4("📋 데이터 테이블", className="mb-3"),
            dash_table.DataTable(
                id='data-table',
                columns=[],
                data=[],
                page_size=20,
                sort_action='native',
                filter_action='native',
                style_table={'overflowX': 'auto'},
                style_cell={
                    'textAlign': 'left',
                    'padding': '10px',
                    'fontSize': '14px'
                },
                style_header={
                    'backgroundColor': '#007bff',
                    'color': 'white',
                    'fontWeight': 'bold'
                },
                style_data_conditional=[
                    {
                        'if': {'row_index': 'odd'},
                        'backgroundColor': '#f8f9fa'
                    }
                ]
            )
        ])
    ], className='mb-4'),

    # 차트 영역
    dbc.Row(id='charts-container', className='mb-4'),

    # 데이터 저장 (숨김)
    dcc.Store(id='stored-data'),
    dcc.Store(id='stored-weeks')

], fluid=True, style={'padding': '20px'})

# ============================================================================
# 콜백: 파일 업로드 처리
# ============================================================================

@app.callback(
    [Output('stored-data', 'data'),
     Output('stored-weeks', 'data'),
     Output('upload-status', 'children'),
     Output('week-dropdown', 'options'),
     Output('compare-week-dropdown', 'options'),
     Output('column-checklist', 'options'),
     Output('column-checklist', 'value'),
     Output('opinion-alert', 'children')],
    Input('upload-data', 'contents'),
    State('upload-data', 'filename')
)
def upload_file(contents, filename):
    if contents is None:
        return None, None, "", [], [], [], [], "📢 의견 알림: 파일을 업로드하면 의견 내용이 여기에 표시됩니다."

    try:
        # 파일 파싱
        df, errors = read_excel_with_week_improved(contents)

        if df is None or df.empty:
            error_msg = "\n".join(errors) if errors else "알 수 없는 오류"
            return None, None, dbc.Alert(f"❌ 파일 읽기 실패: {error_msg}", color="danger"), [], [], [], [], "❌ 파일을 읽을 수 없습니다."

        # 주차 목록 추출
        weeks = sorted(df['주차(B1)'].unique().tolist())
        week_options = [{'label': w, 'value': w} for w in weeks]

        # 컬럼 목록 추출 (숫자 컬럼 우선)
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        all_cols = [col for col in df.columns if col not in ['시트명', '주차(B1)']]
        column_options = [{'label': col, 'value': col} for col in all_cols]
        default_columns = [col for col in ['상품명', '매출', '이익액', '이익률(ROI)'] if col in all_cols]

        # 의견 컬럼 내용 추출
        opinion_text = "📢 의견 알림: "
        if '의견' in df.columns:
            opinions = df['의견'].dropna().unique().tolist()
            if opinions:
                opinion_text += " | ".join([str(op) for op in opinions[:5]])  # 최대 5개
            else:
                opinion_text += "의견 데이터가 없습니다."
        else:
            opinion_text += "'의견' 컬럼이 없습니다."

        success_msg = dbc.Alert(
            f"✅ {filename} 업로드 완료! ({len(df):,}개 행, {len(weeks)}개 주차)",
            color="success"
        )

        return df.to_dict('records'), weeks, success_msg, week_options, week_options, column_options, default_columns, opinion_text

    except Exception as e:
        return None, None, dbc.Alert(f"❌ 오류: {str(e)}", color="danger"), [], [], [], [], "❌ 파일 처리 중 오류 발생"

# ============================================================================
# 콜백: KPI 카드 업데이트
# ============================================================================

@app.callback(
    Output('kpi-cards', 'children'),
    [Input('stored-data', 'data'),
     Input('week-dropdown', 'value')]
)
def update_kpi_cards(data, selected_week):
    if not data or not selected_week:
        return []

    df = pd.DataFrame(data)
    df_filtered = df[df['주차(B1)'] == selected_week].copy()

    if df_filtered.empty:
        return []

    # KPI 계산
    total_sales = df_filtered['매출'].sum() if '매출' in df_filtered.columns else 0
    total_profit = df_filtered['이익액'].sum() if '이익액' in df_filtered.columns else 0
    avg_roi = df_filtered['이익률(ROI)'].mean() if '이익률(ROI)' in df_filtered.columns else 0
    product_count = len(df_filtered)

    cards = [
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5("총 매출", className="card-title"),
                    html.H3(f"₩{total_sales/1000000:.1f}M", className="text-primary")
                ])
            ])
        ], width=3),
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5("총 이익", className="card-title"),
                    html.H3(f"₩{total_profit/1000000:.1f}M", className="text-success")
                ])
            ])
        ], width=3),
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5("평균 ROI", className="card-title"),
                    html.H3(f"{avg_roi:.1f}%", className="text-info")
                ])
            ])
        ], width=3),
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5("상품 수", className="card-title"),
                    html.H3(f"{product_count}개", className="text-warning")
                ])
            ])
        ], width=3)
    ]

    return cards

# ============================================================================
# 콜백: 데이터 테이블 업데이트
# ============================================================================

@app.callback(
    [Output('data-table', 'columns'),
     Output('data-table', 'data')],
    [Input('stored-data', 'data'),
     Input('week-dropdown', 'value'),
     Input('column-checklist', 'value')]
)
def update_table(data, selected_week, selected_columns):
    if not data or not selected_week or not selected_columns:
        return [], []

    df = pd.DataFrame(data)
    df_filtered = df[df['주차(B1)'] == selected_week].copy()

    # 선택된 컬럼만 표시
    display_cols = [col for col in selected_columns if col in df_filtered.columns]
    df_display = df_filtered[display_cols]

    columns = [{'name': col, 'id': col} for col in display_cols]
    data_records = df_display.to_dict('records')

    return columns, data_records

# ============================================================================
# 콜백: 차트 생성
# ============================================================================

@app.callback(
    Output('charts-container', 'children'),
    [Input('stored-data', 'data'),
     Input('week-dropdown', 'value'),
     Input('compare-week-dropdown', 'value'),
     Input('column-checklist', 'value')]
)
def update_charts(data, selected_week, compare_week, selected_columns):
    if not data or not selected_week or not selected_columns:
        return []

    df = pd.DataFrame(data)
    df_filtered = df[df['주차(B1)'] == selected_week].copy()

    # 숫자 컬럼만 차트 생성
    numeric_cols = [col for col in selected_columns if col in df_filtered.select_dtypes(include=['number']).columns]

    charts = []

    for col in numeric_cols:
        if col in df_filtered.columns and not df_filtered[col].isna().all():
            # Top 20 막대 차트
            df_sorted = df_filtered.nlargest(20, col)
            fig_bar = px.bar(
                df_sorted,
                x='상품명' if '상품명' in df_sorted.columns else df_sorted.index,
                y=col,
                title=f"{col} - Top 20",
                labels={'x': '상품명', 'y': col}
            )
            fig_bar.update_layout(height=400, xaxis_tickangle=-45)

            # 히스토그램
            fig_hist = px.histogram(
                df_filtered,
                x=col,
                title=f"{col} - 분포",
                nbins=30
            )
            fig_hist.update_layout(height=400)

            charts.append(
                dbc.Row([
                    dbc.Col([dcc.Graph(figure=fig_bar)], width=6),
                    dbc.Col([dcc.Graph(figure=fig_hist)], width=6)
                ], className='mb-4')
            )

    return charts

# ============================================================================
# 서버 실행
# ============================================================================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8050))
    print(f"\n{'='*60}")
    print("🚀 고급 트래픽 분석 대시보드 시작")
    print(f"{'='*60}")
    print(f"📍 포트: {port}")
    print(f"🌐 URL: http://0.0.0.0:{port}")
    print(f"{'='*60}\n")

    app.run_server(host='0.0.0.0', port=port, debug=False)
