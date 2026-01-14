
"""
고급 트래픽 데이터 분석 대시보드
- 주차별 필터링 및 비교
- 동적 컬럼 선택
- 인터랙티브 차트 생성
- M1 의견 강조 표시
"""

import dash
from dash import dcc, html, Input, Output, State, dash_table, ctx
import dash_bootstrap_components as dbc
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import numpy as np
import base64
from io import BytesIO
import openpyxl
from datetime import datetime

# Initialize Dash app with Bootstrap theme
app = dash.Dash(
    __name__,
    external_stylesheets=[dbc.themes.BOOTSTRAP, dbc.icons.FONT_AWESOME],
    suppress_callback_exceptions=True
)

server = app.server

# ==================== DATA PARSING FUNCTIONS ====================

def parse_excel_with_week_info(contents):
    """
    Parse Excel file with B1 week format
    Returns: dict with success, data, week_info, sheets, columns
    """
    try:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)

        wb = openpyxl.load_workbook(BytesIO(decoded))

        all_data = []
        week_info = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Extract week from B1
            week_b1 = ws['B1'].value
            if week_b1:
                week_info[sheet_name] = str(week_b1).strip()
            else:
                week_info[sheet_name] = sheet_name

            # Get headers from row 2
            headers = []
            for cell in ws[2]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            if not headers:
                continue

            # Get data rows starting from row 3
            data = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not any(row):
                    continue
                row_data = list(row[:len(headers)])
                data.append(row_data)

            if data:
                df_sheet = pd.DataFrame(data, columns=headers)
                df_sheet['시트명'] = sheet_name
                df_sheet['주차(B1)'] = week_info[sheet_name]
                all_data.append(df_sheet)

        if all_data:
            df_combined = pd.concat(all_data, ignore_index=True)

            # Identify numeric columns
            numeric_cols = []
            for col in df_combined.columns:
                if col not in ['상품명', '주차', '쇼핑몰', '의견', '시트명', '주차(B1)']:
                    try:
                        pd.to_numeric(df_combined[col], errors='coerce')
                        numeric_cols.append(col)
                    except:
                        pass

            return {
                'success': True,
                'data': df_combined,
                'week_info': week_info,
                'sheets': list(wb.sheetnames),
                'columns': list(df_combined.columns),
                'numeric_columns': numeric_cols
            }
        else:
            return {'success': False, 'error': '데이터를 찾을 수 없습니다.'}

    except Exception as e:
        return {'success': False, 'error': f'파일 파싱 오류: {str(e)}'}


def identify_column_types(df):
    """Identify which columns are numeric for charting"""
    numeric_cols = []
    text_cols = []

    for col in df.columns:
        if col in ['시트명', '주차(B1)']:
            continue
        try:
            # Try to convert to numeric
            test = pd.to_numeric(df[col], errors='coerce')
            if test.notna().sum() > len(df) * 0.5:  # If >50% are numeric
                numeric_cols.append(col)
            else:
                text_cols.append(col)
        except:
            text_cols.append(col)

    return numeric_cols, text_cols


# ==================== LAYOUT ====================

app.layout = dbc.Container([
    # Header
    dbc.Row([
        dbc.Col([
            html.H1([
                html.I(className="fas fa-chart-line me-3"),
                "고급 트래픽 데이터 분석 대시보드"
            ], className="text-center text-primary mb-4 mt-4")
        ])
    ]),

    # M1 Opinion Alert Box (initially hidden)
    dbc.Row([
        dbc.Col([
            dbc.Alert(
                id="m1-opinion-box",
                children="",
                color="info",
                className="mb-4",
                is_open=False,
                dismissable=True
            )
        ])
    ]),

    # File Upload Section
    dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H4([
                        html.I(className="fas fa-upload me-2"),
                        "Excel 파일 업로드"
                    ], className="card-title"),
                    dcc.Upload(
                        id='upload-data',
                        children=html.Div([
                            html.I(className="fas fa-cloud-upload-alt fa-3x mb-3"),
                            html.P('Excel 파일을 드래그하거나 클릭하여 업로드하세요', className="mb-1"),
                            html.P('지원 형식: .xlsx, .xls', className="text-muted small")
                        ]),
                        style={
                            'width': '100%',
                            'height': '150px',
                            'lineHeight': '150px',
                            'borderWidth': '2px',
                            'borderStyle': 'dashed',
                            'borderRadius': '10px',
                            'textAlign': 'center',
                            'backgroundColor': '#f8f9fa'
                        },
                        multiple=False
                    ),
                    html.Div(id='upload-status', className="mt-3")
                ])
            ], className="shadow-sm mb-4")
        ])
    ]),

    # Filter Control Panel
    dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5([
                        html.I(className="fas fa-filter me-2"),
                        "필터 및 옵션"
                    ], className="card-title"),
                    dbc.Row([
                        # Week selection
                        dbc.Col([
                            html.Label("📅 주차 선택", className="fw-bold"),
                            dcc.Dropdown(
                                id='week-selector',
                                placeholder='주차를 선택하세요...',
                                className="mb-3"
                            )
                        ], md=4),

                        # Comparison week selection
                        dbc.Col([
                            html.Label("🔄 비교 주차 선택", className="fw-bold"),
                            dcc.Dropdown(
                                id='comparison-week-selector',
                                placeholder='비교할 주차를 선택하세요...',
                                className="mb-3"
                            )
                        ], md=4),

                        # Column visibility selection
                        dbc.Col([
                            html.Label("👁️ 표시할 컬럼 선택", className="fw-bold"),
                            dcc.Dropdown(
                                id='column-selector',
                                multi=True,
                                placeholder='컬럼을 선택하세요...',
                                className="mb-3"
                            )
                        ], md=4),
                    ])
                ])
            ], className="shadow-sm mb-4")
        ], id="filter-panel", style={'display': 'none'})
    ]),

    # KPI Cards
    dbc.Row(id='kpi-cards', className="mb-4"),

    # Comparison Results (when comparison is active)
    dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardBody(id='comparison-results')
            ], className="shadow-sm mb-4")
        ], id="comparison-panel", style={'display': 'none'})
    ]),

    # Data Table
    dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader([
                    html.H5([
                        html.I(className="fas fa-table me-2"),
                        "데이터 테이블"
                    ], className="mb-0")
                ]),
                dbc.CardBody([
                    dash_table.DataTable(
                        id='data-table',
                        style_table={'overflowX': 'auto'},
                        style_cell={
                            'textAlign': 'left',
                            'padding': '10px',
                            'fontFamily': 'Arial, sans-serif'
                        },
                        style_header={
                            'backgroundColor': '#0d6efd',
                            'color': 'white',
                            'fontWeight': 'bold',
                            'textAlign': 'center'
                        },
                        style_data_conditional=[
                            {
                                'if': {'row_index': 'odd'},
                                'backgroundColor': '#f8f9fa'
                            }
                        ],
                        sort_action='native',
                        filter_action='native',
                        page_action='native',
                        page_size=20,
                        page_current=0
                    )
                ])
            ], className="shadow-sm mb-4")
        ], id="table-panel", style={'display': 'none'})
    ]),

    # Dynamic Charts Grid
    dbc.Row(id='charts-grid', className="mb-4"),

    # Hidden data store
    dcc.Store(id='stored-data'),
    dcc.Store(id='stored-parsed-result')

], fluid=True, style={'backgroundColor': '#f0f2f5', 'minHeight': '100vh', 'paddingBottom': '50px'})


# ==================== CALLBACKS ====================

@app.callback(
    [Output('stored-parsed-result', 'data'),
     Output('upload-status', 'children'),
     Output('filter-panel', 'style'),
     Output('table-panel', 'style'),
     Output('m1-opinion-box', 'children'),
     Output('m1-opinion-box', 'is_open'),
     Output('week-selector', 'options'),
     Output('week-selector', 'value'),
     Output('comparison-week-selector', 'options'),
     Output('column-selector', 'options'),
     Output('column-selector', 'value')],
    [Input('upload-data', 'contents')],
    [State('upload-data', 'filename')]
)
def handle_upload(contents, filename):
    """Handle file upload and initial processing"""
    if contents is None:
        return (None, "", {'display': 'none'}, {'display': 'none'}, 
                "", False, [], None, [], [], [])

    # Parse the file
    result = parse_excel_with_week_info(contents)

    if not result['success']:
        error_msg = dbc.Alert([
            html.I(className="fas fa-exclamation-triangle me-2"),
            f"오류: {result['error']}"
        ], color="danger")
        return (None, error_msg, {'display': 'none'}, {'display': 'none'},
                "", False, [], None, [], [], [])

    df = result['data']

    # Success message
    success_msg = dbc.Alert([
        html.I(className="fas fa-check-circle me-2"),
        f"✓ {filename} 업로드 완료! ({len(df)} 개 레코드)"
    ], color="success")

    # Get unique weeks
    weeks = sorted(df['주차(B1)'].unique().tolist())
    week_options = [{'label': w, 'value': w} for w in weeks]

    # Get all columns for selection
    exclude_cols = ['시트명', '주차(B1)']
    available_cols = [col for col in df.columns if col not in exclude_cols]
    col_options = [{'label': col, 'value': col} for col in available_cols]

    # Default selected columns (all numeric + key text columns)
    default_cols = ['상품명', '주차', '쇼핑몰'] + result.get('numeric_columns', [])
    default_cols = [c for c in default_cols if c in available_cols]

    # Get M1 opinions (assuming '의견' column exists)
    m1_opinions = ""
    if '의견' in df.columns:
        unique_opinions = df['의견'].dropna().unique()
        if len(unique_opinions) > 0:
            m1_opinions = html.Div([
                html.H5([
                    html.I(className="fas fa-comments me-2"),
                    "주요 의견"
                ]),
                html.Ul([html.Li(op) for op in unique_opinions[:5]])
            ])

    return (result, success_msg, {'display': 'block'}, {'display': 'block'},
            m1_opinions, bool(m1_opinions), week_options, weeks[0] if weeks else None,
            week_options, col_options, default_cols)


@app.callback(
    [Output('data-table', 'data'),
     Output('data-table', 'columns'),
     Output('kpi-cards', 'children'),
     Output('comparison-panel', 'style'),
     Output('comparison-results', 'children')],
    [Input('week-selector', 'value'),
     Input('comparison-week-selector', 'value'),
     Input('column-selector', 'value'),
     Input('stored-parsed-result', 'data')]
)
def update_table_and_kpis(selected_week, comparison_week, selected_columns, parsed_result):
    """Update data table and KPI cards based on filters"""
    if not parsed_result or not selected_week:
        return [], [], [], {'display': 'none'}, ""

    df = pd.DataFrame(parsed_result['data'])

    # Filter by selected week
    df_filtered = df[df['주차(B1)'] == selected_week].copy()

    # Apply column selection
    if selected_columns:
        display_cols = [col for col in selected_columns if col in df_filtered.columns]
        df_display = df_filtered[display_cols]
    else:
        df_display = df_filtered

    # Prepare table data
    table_columns = [{"name": col, "id": col} for col in df_display.columns]
    table_data = df_display.to_dict('records')

    # Calculate KPIs
    kpi_cards = create_kpi_cards(df_filtered, parsed_result.get('numeric_columns', []))

    # Handle comparison if comparison week is selected
    comparison_content = ""
    comparison_style = {'display': 'none'}

    if comparison_week and comparison_week != selected_week:
        df_comparison = df[df['주차(B1)'] == comparison_week].copy()
        comparison_content = create_comparison_view(df_filtered, df_comparison, 
                                                     selected_week, comparison_week,
                                                     parsed_result.get('numeric_columns', []))
        comparison_style = {'display': 'block'}

    return table_data, table_columns, kpi_cards, comparison_style, comparison_content


@app.callback(
    Output('charts-grid', 'children'),
    [Input('week-selector', 'value'),
     Input('column-selector', 'value'),
     Input('stored-parsed-result', 'data')]
)
def update_charts(selected_week, selected_columns, parsed_result):
    """Generate dynamic charts for selected numeric columns"""
    if not parsed_result or not selected_week or not selected_columns:
        return []

    df = pd.DataFrame(parsed_result['data'])
    df_filtered = df[df['주차(B1)'] == selected_week].copy()

    # Identify numeric columns from selection
    numeric_cols = [col for col in selected_columns 
                   if col in parsed_result.get('numeric_columns', [])]

    if not numeric_cols:
        return [dbc.Alert("선택된 숫자 컬럼이 없습니다. 차트를 표시하려면 숫자 데이터가 포함된 컬럼을 선택하세요.", 
                         color="info")]

    charts = []

    for col in numeric_cols:
        try:
            # Convert to numeric
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce')
            df_chart = df_filtered.dropna(subset=[col])

            if len(df_chart) == 0:
                continue

            # Create multiple chart types for each column
            charts.append(create_column_charts(df_chart, col))

        except Exception as e:
            print(f"Error creating chart for {col}: {e}")
            continue

    return charts


def create_kpi_cards(df, numeric_columns):
    """Create KPI cards for key metrics"""
    cards = []

    # Define KPI calculations
    kpis = []

    if '매출' in df.columns:
        total_sales = pd.to_numeric(df['매출'], errors='coerce').sum()
        kpis.append({
            'icon': 'fa-won-sign',
            'title': '총 매출',
            'value': f'₩{total_sales:,.0f}',
            'color': 'primary'
        })

    if '이익액' in df.columns:
        total_profit = pd.to_numeric(df['이익액'], errors='coerce').sum()
        kpis.append({
            'icon': 'fa-chart-line',
            'title': '총 이익',
            'value': f'₩{total_profit:,.0f}',
            'color': 'success'
        })

    if '이익률(ROI)' in df.columns:
        avg_roi = pd.to_numeric(df['이익률(ROI)'], errors='coerce').mean()
        kpis.append({
            'icon': 'fa-percent',
            'title': '평균 이익률',
            'value': f'{avg_roi:.2f}%',
            'color': 'info'
        })

    if '트래픽 비용' in df.columns:
        total_traffic_cost = pd.to_numeric(df['트래픽 비용'], errors='coerce').sum()
        kpis.append({
            'icon': 'fa-credit-card',
            'title': '총 트래픽 비용',
            'value': f'₩{total_traffic_cost:,.0f}',
            'color': 'warning'
        })

    # Add product count
    kpis.append({
        'icon': 'fa-box',
        'title': '상품 수',
        'value': str(len(df)),
        'color': 'secondary'
    })

    # Create card components
    for kpi in kpis:
        card = dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.Div([
                        html.I(className=f"fas {kpi['icon']} fa-2x mb-2 text-{kpi['color']}"),
                        html.H6(kpi['title'], className="text-muted mb-1"),
                        html.H3(kpi['value'], className=f"text-{kpi['color']} fw-bold")
                    ], className="text-center")
                ])
            ], className="shadow-sm h-100")
        ], md=6, lg=2, className="mb-3")
        cards.append(card)

    return cards


def create_column_charts(df, column_name):
    """Create comprehensive charts for a specific column"""

    # Sort by column value and take top 20
    df_sorted = df.sort_values(column_name, ascending=False).head(20)

    # Create bar chart
    fig_bar = px.bar(
        df_sorted,
        x='상품명' if '상품명' in df.columns else df_sorted.index,
        y=column_name,
        title=f'{column_name} - 상위 20개',
        labels={column_name: column_name, '상품명': '상품'},
        color=column_name,
        color_continuous_scale='Blues'
    )
    fig_bar.update_layout(
        xaxis_tickangle=-45,
        height=400,
        showlegend=False
    )

    # Create distribution histogram
    fig_hist = px.histogram(
        df,
        x=column_name,
        nbins=20,
        title=f'{column_name} - 분포',
        labels={column_name: column_name}
    )
    fig_hist.update_layout(height=400)

    return dbc.Col([
        dbc.Card([
            dbc.CardHeader(html.H5([
                html.I(className="fas fa-chart-bar me-2"),
                column_name
            ])),
            dbc.CardBody([
                dbc.Row([
                    dbc.Col([
                        dcc.Graph(figure=fig_bar)
                    ], md=6),
                    dbc.Col([
                        dcc.Graph(figure=fig_hist)
                    ], md=6)
                ])
            ])
        ], className="shadow-sm mb-4")
    ], md=12)


def create_comparison_view(df1, df2, week1, week2, numeric_columns):
    """Create comparison view between two weeks"""

    comparisons = []

    for col in numeric_columns:
        if col in df1.columns and col in df2.columns:
            try:
                val1 = pd.to_numeric(df1[col], errors='coerce').sum()
                val2 = pd.to_numeric(df2[col], errors='coerce').sum()

                if val1 > 0:
                    change_pct = ((val2 - val1) / val1) * 100
                    change_color = 'success' if change_pct > 0 else 'danger'
                    change_icon = 'fa-arrow-up' if change_pct > 0 else 'fa-arrow-down'

                    comparisons.append(
                        dbc.Col([
                            dbc.Card([
                                dbc.CardBody([
                                    html.H6(col, className="text-muted"),
                                    html.Div([
                                        html.Span(f"{week1}: ", className="fw-bold"),
                                        html.Span(f"₩{val1:,.0f}")
                                    ]),
                                    html.Div([
                                        html.Span(f"{week2}: ", className="fw-bold"),
                                        html.Span(f"₩{val2:,.0f}")
                                    ]),
                                    html.Div([
                                        html.I(className=f"fas {change_icon} me-1 text-{change_color}"),
                                        html.Span(f"{change_pct:+.2f}%", 
                                                className=f"text-{change_color} fw-bold")
                                    ], className="mt-2")
                                ])
                            ], className="shadow-sm h-100")
                        ], md=4, lg=3, className="mb-3")
                    )
            except:
                continue

    if comparisons:
        return html.Div([
            html.H5([
                html.I(className="fas fa-exchange-alt me-2"),
                f"주차 비교: {week1} vs {week2}"
            ], className="mb-3"),
            dbc.Row(comparisons)
        ])
    else:
        return dbc.Alert("비교 가능한 데이터가 없습니다.", color="warning")


# ==================== RUN SERVER ====================

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 8050))
    app.run_server(host='0.0.0.0', port=port, debug=False)
